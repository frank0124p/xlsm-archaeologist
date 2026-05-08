"""Orchestrate extraction + analysis: open workbook → extract → analyze → serialize."""

from __future__ import annotations

from pathlib import Path

from xlsm_archaeologist.analyzers.dependency_analyzer import graph_to_json, run_dependency_analysis
from xlsm_archaeologist.analyzers.formula_analyzer import analyze_formulas
from xlsm_archaeologist.analyzers.vba_analyzer import analyze_vba
from xlsm_archaeologist.extractors.cell_extractor import extract_cells
from xlsm_archaeologist.extractors.named_range_extractor import extract_named_ranges
from xlsm_archaeologist.extractors.sheet_extractor import extract_sheets
from xlsm_archaeologist.extractors.validation_extractor import extract_validations
from xlsm_archaeologist.extractors.workbook_extractor import extract_workbook
from xlsm_archaeologist.reports.architecture_report import build_architecture_md
from xlsm_archaeologist.reports.cross_sheet_refs_report import build_cross_sheet_refs
from xlsm_archaeologist.reports.data_flow_report import build_data_flow_md
from xlsm_archaeologist.reports.formula_categories_report import build_categories_report
from xlsm_archaeologist.reports.hotspot_cells_report import build_hotspot_cells
from xlsm_archaeologist.reports.integration_report import build_integration_md
from xlsm_archaeologist.reports.summary_builder import build_summary
from xlsm_archaeologist.reports.top_complex_formulas_report import build_top_complex_formulas
from xlsm_archaeologist.reports.vba_behavior_report import build_vba_behavior
from xlsm_archaeologist.serializers.csv_writer import write_csv
from xlsm_archaeologist.serializers.json_writer import write_json
from xlsm_archaeologist.utils.logging import get_logger
from xlsm_archaeologist.utils.progress import ProgressBar

logger = get_logger(__name__)

# Column order must match DATA_MODEL.md / reference/csv_schemas.md exactly
_SHEETS_COLUMNS = [
    "sheet_id",
    "sheet_name",
    "sheet_index",
    "is_hidden",
    "is_very_hidden",
    "used_range",
    "row_count",
    "col_count",
    "cell_count",
    "formula_cell_count",
]

_NAMED_RANGES_COLUMNS = [
    "named_range_id",
    "range_name",
    "scope",
    "refers_to",
    "has_dynamic_formula",
    "is_valid",
]

_CELLS_COLUMNS = [
    "cell_id",
    "sheet_name",
    "cell_address",
    "qualified_address",
    "cell_row",
    "cell_col",
    "has_formula",
    "has_validation",
    "is_named",
    "is_referenced",
    "value_type",
    "raw_value",
]

_VALIDATIONS_COLUMNS = [
    "validation_id",
    "qualified_address",
    "range_text",
    "validation_type",
    "formula1",
    "formula2",
    "enum_values",
    "allow_blank",
    "error_title",
    "error_message",
]

_DEPENDENCIES_COLUMNS = [
    "dependency_id",
    "source_qualified_address",
    "target_qualified_address",
    "via",
    "via_detail",
    "is_cross_sheet",
]


def _formula_dict(f: object) -> dict[str, object]:
    """Serialize a FormulaRecord to a plain dict, converting CellRef list to dicts."""
    from xlsm_archaeologist.models.formula import FormulaRecord

    if not isinstance(f, FormulaRecord):
        return {}
    d: dict[str, object] = f.model_dump()
    d["referenced_cells"] = [{"sheet": r.sheet, "address": r.address} for r in f.referenced_cells]
    return d


def _build_named_addresses(wb: object) -> set[str]:
    """Return the set of qualified addresses targeted by all named ranges."""
    from openpyxl.utils.cell import range_boundaries
    from openpyxl.workbook.workbook import Workbook

    if not isinstance(wb, Workbook):
        return set()

    addresses: set[str] = set()
    for name in wb.defined_names:
        if name.startswith("_xlnm."):
            continue
        dn = wb.defined_names[name]
        raw: str = getattr(dn, "attr_text", None) or getattr(dn, "value", "") or ""
        ref = raw.lstrip("=").strip()
        if "!" not in ref:
            continue
        try:
            sheet_part, range_part = ref.split("!", 1)
            sheet_name = sheet_part.strip("'$")
            rp = range_part.replace("$", "")
            if ":" not in rp:
                # Single cell
                addresses.add(f"{sheet_name}!{rp}")
            else:
                min_col, min_row, max_col, max_row = range_boundaries(rp)
                from openpyxl.utils import get_column_letter

                if min_col is None or min_row is None or max_col is None or max_row is None:
                    continue
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        addresses.add(f"{sheet_name}!{get_column_letter(c)}{r}")
        except Exception:
            pass
    return addresses


def _build_validation_addresses(wb: object) -> set[str]:
    """Return the set of qualified addresses covered by any data validation."""
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import range_boundaries
    from openpyxl.workbook.workbook import Workbook

    if not isinstance(wb, Workbook):
        return set()

    addresses: set[str] = set()
    for sheet in wb.worksheets:
        for dv in sheet.data_validations.dataValidation:
            sqref = str(dv.sqref) if dv.sqref else ""
            for token in sqref.split():
                token = token.replace("$", "")
                if ":" in token:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(token)
                        if min_col is None or min_row is None or max_col is None or max_row is None:
                            continue
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                addresses.add(f"{sheet.title}!{get_column_letter(c)}{r}")
                    except Exception:
                        pass
                else:
                    addresses.add(f"{sheet.title}!{token}")
    return addresses


def run_extraction(
    input_path: Path,
    output_dir: Path,
    quiet: bool = False,
    log_level: str = "info",
) -> None:
    """Run extraction + formula analysis and write output files 01-07.

    Args:
        input_path: Path to the source .xlsm file.
        output_dir: Directory to write all output files into.
        quiet: When True, suppress the progress bar.
        log_level: Logging verbosity level string.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    all_warnings: list[str] = []

    with ProgressBar(quiet=quiet) as bar:
        task = bar.add_task("Extraction + Analysis", total=10)

        # --- Step 1: Workbook metadata ---
        workbook_record, wb = extract_workbook(input_path)
        bar.advance(task)

        # --- Step 2: Sheets ---
        sheets = sorted(extract_sheets(wb), key=lambda s: s.sheet_index)
        bar.advance(task)

        # --- Step 3: Named ranges ---
        named_ranges = sorted(extract_named_ranges(wb), key=lambda n: n.range_name)
        named_addresses = _build_named_addresses(wb)
        bar.advance(task)

        # --- Step 4: Validations ---
        validations = sorted(extract_validations(wb), key=lambda v: v.qualified_address)
        validation_addresses = _build_validation_addresses(wb)
        bar.advance(task)

        # --- Step 5: Cells ---
        cells = sorted(
            extract_cells(wb, named_addresses, validation_addresses),
            key=lambda c: c.qualified_address,
        )
        # Re-number cell_ids after sort (order matches qualified_address sort)
        cells = [c.model_copy(update={"cell_id": i + 1}) for i, c in enumerate(cells)]
        bar.advance(task)

        # --- Step 5b: Formula analysis ---
        formula_warnings: list[str] = []
        formulas = sorted(
            analyze_formulas(cells, formula_warnings),
            key=lambda f: f.qualified_address,
        )
        formulas = [f.model_copy(update={"formula_id": i + 1}) for i, f in enumerate(formulas)]
        all_warnings.extend(formula_warnings)
        bar.advance(task)

        # --- Step 5c: VBA analysis ---
        vba_warnings: list[str] = []
        vba_modules, vba_procedures = analyze_vba(input_path, vba_warnings)
        all_warnings.extend(vba_warnings)
        bar.advance(task)

        # --- Step 5d: Dependency graph ---
        dep_warnings: list[str] = []
        _graph, dep_edges, cycles, orphan_ids, cells = run_dependency_analysis(
            formulas, vba_procedures, named_ranges, cells, validations, dep_warnings
        )
        all_warnings.extend(dep_warnings)
        bar.advance(task)

        # --- Step 6: Write files ---
        write_json(
            output_dir / "01_workbook.json",
            {"workbook": workbook_record.model_dump()},
        )

        write_csv(
            output_dir / "02_sheets.csv",
            [s.model_dump() for s in sheets],
            _SHEETS_COLUMNS,
        )

        write_csv(
            output_dir / "03_named_ranges.csv",
            [n.model_dump() for n in named_ranges],
            _NAMED_RANGES_COLUMNS,
        )

        write_csv(
            output_dir / "04_cells.csv",
            [c.model_dump() for c in cells],
            _CELLS_COLUMNS,
        )

        write_json(
            output_dir / "05_formulas.json",
            {
                "formulas": [_formula_dict(f) for f in formulas],
                "warnings": formula_warnings,
            },
        )

        write_csv(
            output_dir / "06_validations.csv",
            [v.model_dump() for v in validations],
            _VALIDATIONS_COLUMNS,
        )

        write_json(
            output_dir / "07_vba_modules.json",
            {"vba_modules": [m.model_dump() for m in vba_modules]},
        )

        write_json(
            output_dir / "08_vba_procedures.json",
            {"vba_procedures": [p.model_dump() for p in vba_procedures]},
        )

        write_csv(
            output_dir / "09_dependencies.csv",
            [e.model_dump() for e in dep_edges],
            _DEPENDENCIES_COLUMNS,
        )

        write_json(
            output_dir / "10_dependency_graph.json",
            graph_to_json(_graph, cycles),
        )

        # Rewrite 04_cells.csv with backfilled is_referenced
        write_csv(
            output_dir / "04_cells.csv",
            [c.model_dump() for c in cells],
            _CELLS_COLUMNS,
        )

        reports_dir = output_dir / "reports"
        reports_dir.mkdir(parents=True, exist_ok=True)

        write_json(
            reports_dir / "cycles.json",
            {"cycles": [c.model_dump() for c in cycles]},
        )

        bar.advance(task)

        # --- Step 9: Phase 6 reports ---
        from xlsm_archaeologist import __version__

        summary = build_summary(
            input_path=input_path,
            tool_version=__version__,
            sheets=sheets,
            named_ranges=named_ranges,
            formulas=formulas,
            validations=validations,
            vba_modules=vba_modules,
            vba_procedures=vba_procedures,
            dep_edges=dep_edges,
            cycles=cycles,
            orphan_ids=orphan_ids,
            raw_warnings=all_warnings,
        )

        write_json(output_dir / "00_summary.json", summary.model_dump())

        write_csv(
            reports_dir / "formula_categories.csv",
            build_categories_report(formulas),
            ["category", "formula_count", "total_complexity", "avg_complexity",
             "max_complexity", "pct_of_total"],
        )

        write_csv(
            reports_dir / "top_complex_formulas.csv",
            build_top_complex_formulas(formulas),
            ["rank", "qualified_address", "formula_text", "formula_category",
             "nesting_depth", "function_count", "referenced_cell_count", "complexity_score"],
        )

        write_csv(
            reports_dir / "hotspot_cells.csv",
            build_hotspot_cells(_graph, cells),
            ["rank", "qualified_address", "node_type", "in_degree",
             "referenced_by_formula_count", "referenced_by_vba_count", "value_type", "raw_value"],
        )

        write_csv(
            reports_dir / "vba_behavior.csv",
            build_vba_behavior(vba_procedures, vba_modules),
            ["module_name", "procedure_name", "procedure_type", "line_count",
             "read_count", "write_count", "cross_sheet_read_count", "cross_sheet_write_count",
             "has_dynamic_range", "has_event_trigger", "call_count", "complexity_score"],
        )

        write_csv(
            reports_dir / "cross_sheet_refs.csv",
            build_cross_sheet_refs(dep_edges),
            ["source_qualified_address", "source_sheet", "target_qualified_address",
             "target_sheet", "via", "via_detail"],
        )

        # Developer documentation reports
        source_name = input_path.name
        (reports_dir / "architecture.md").write_text(
            build_architecture_md(sheets, dep_edges, formulas, vba_modules, source_name),
            encoding="utf-8",
        )
        (reports_dir / "data_flow.md").write_text(
            build_data_flow_md(sheets, cells, formulas, validations, dep_edges, vba_procedures, source_name, vba_modules),  # noqa: E501
            encoding="utf-8",
        )
        (reports_dir / "integration.md").write_text(
            build_integration_md(source_name, summary, sheets, validations, formulas, str(output_dir)),  # noqa: E501
            encoding="utf-8",
        )

        bar.advance(task)

    wb.close()
    logger.info(
        "Complete → %s  score=%d difficulty=%s  (%d sheets, %d formulas, %d warnings)",
        output_dir,
        summary.complexity_score,
        summary.migration_difficulty,
        len(sheets),
        len(formulas),
        len(all_warnings),
    )
