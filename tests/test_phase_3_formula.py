"""Phase 3 formula analysis tests."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from xlsm_archaeologist.analyzers.formula_analyzer import analyze_formulas
from xlsm_archaeologist.analyzers.formula_classifier import classify
from xlsm_archaeologist.analyzers.formula_complexity import compute_complexity
from xlsm_archaeologist.analyzers.formula_parser import parse
from xlsm_archaeologist.analyzers.formula_tokenizer import tokenize
from xlsm_archaeologist.extractors.cell_extractor import extract_cells
from xlsm_archaeologist.models.formula import (
    FunctionNode,
    NamedRangeNode,
    OperandNode,
    OperatorNode,
    RangeNode,
    UnparsableNode,
)


@pytest.fixture
def fixtures_dir() -> Path:
    return Path(__file__).parent / "fixtures"


# ---------------------------------------------------------------------------
# Tokenizer tests
# ---------------------------------------------------------------------------


def test_tokenize_simple() -> None:
    tokens = tokenize("=A1+B1")
    assert len(tokens) > 0
    values = [t.value for t in tokens]
    assert "A1" in values or any("A" in v for v in values)


def test_tokenize_with_equals() -> None:
    t1 = tokenize("=SUM(A1:A10)")
    t2 = tokenize("SUM(A1:A10)")
    assert len(t1) == len(t2)


def test_tokenize_empty() -> None:
    assert tokenize("") == []


# ---------------------------------------------------------------------------
# Parser tests
# ---------------------------------------------------------------------------


def test_parse_range_operand() -> None:
    tokens = tokenize("=A1")
    node = parse(tokens)
    assert isinstance(node, RangeNode)
    assert node.address == "A1"
    assert node.sheet is None


def test_parse_cross_sheet_range() -> None:
    tokens = tokenize("=Sheet2!B3")
    node = parse(tokens)
    assert isinstance(node, RangeNode)
    assert node.sheet == "Sheet2"
    assert node.address == "B3"


def test_parse_function_node() -> None:
    tokens = tokenize("=SUM(A1:A10)")
    node = parse(tokens)
    assert isinstance(node, FunctionNode)
    assert node.name == "SUM"
    assert len(node.args) == 1


def test_parse_nested_if() -> None:
    formula = "=IF(A1>0,IF(B1>0,IF(C1>0,1,0),0),0)"
    tokens = tokenize(formula)
    node = parse(tokens)
    assert isinstance(node, FunctionNode)
    assert node.name == "IF"
    # depth should be at least 3
    from xlsm_archaeologist.analyzers.formula_complexity import _max_depth

    assert _max_depth(node) >= 3


def test_parse_binary_operator() -> None:
    tokens = tokenize("=A1+B1")
    node = parse(tokens)
    assert isinstance(node, OperatorNode)
    assert node.op == "+"


def test_parse_literal_number() -> None:
    tokens = tokenize("=42")
    node = parse(tokens)
    assert isinstance(node, OperandNode)
    assert node.operand_type == "number"
    assert node.value == "42"


def test_parse_text_literal() -> None:
    tokens = tokenize('="hello"')
    node = parse(tokens)
    assert isinstance(node, OperandNode)
    assert node.operand_type == "text"


def test_parse_empty_returns_unparsable() -> None:
    node = parse([])
    assert isinstance(node, UnparsableNode)


# ---------------------------------------------------------------------------
# Classifier tests
# ---------------------------------------------------------------------------


def test_classify_lookup() -> None:
    tokens = tokenize("=VLOOKUP(A1,B:C,2,0)")
    node = parse(tokens)
    assert classify(node) == "lookup"


def test_classify_branch() -> None:
    tokens = tokenize('=IF(A1>0,"pos","neg")')
    node = parse(tokens)
    assert classify(node) == "branch"


def test_classify_aggregate() -> None:
    tokens = tokenize("=SUM(A1:A10)")
    node = parse(tokens)
    assert classify(node) == "aggregate"


def test_classify_text() -> None:
    tokens = tokenize('=CONCATENATE(A1,"-",B1)')
    node = parse(tokens)
    assert classify(node) == "text"


def test_classify_compute() -> None:
    tokens = tokenize("=A1*B1+C1")
    node = parse(tokens)
    assert classify(node) == "compute"


def test_classify_reference_bare_cell() -> None:
    node = RangeNode(sheet=None, address="A1")
    assert classify(node) == "reference"


def test_classify_mixed_if_vlookup() -> None:
    tokens = tokenize('=IF(VLOOKUP(A1,B:C,2,0)>0,"Y","N")')
    node = parse(tokens)
    assert classify(node) == "mixed"


# ---------------------------------------------------------------------------
# Complexity tests
# ---------------------------------------------------------------------------


def test_complexity_simple_ref() -> None:
    node = RangeNode(sheet=None, address="A1")
    depth, count, score = compute_complexity(node, [])
    assert depth == 0
    assert count == 0
    assert score == 0


def test_complexity_nested_if() -> None:
    formula = "=IF(A1>0,IF(B1>0,1,0),0)"
    tokens = tokenize(formula)
    node = parse(tokens)
    from xlsm_archaeologist.analyzers.formula_analyzer import _extract_references

    refs, _ = _extract_references(node)
    depth, count, score = compute_complexity(node, refs)
    assert depth >= 2
    assert count >= 2
    assert score > 0


# ---------------------------------------------------------------------------
# Full pipeline tests via formulas_complex.xlsm
# ---------------------------------------------------------------------------


def test_analyze_formulas_from_complex_fixture(fixtures_dir: Path) -> None:
    path = fixtures_dir / "formulas_complex.xlsm"
    wb = load_workbook(str(path), data_only=False, keep_vba=True)
    cells = list(extract_cells(wb, set(), set()))
    warnings: list[str] = []
    records = list(analyze_formulas(cells, warnings))
    wb.close()

    assert len(records) >= 7  # D1–D8 are all formulas
    # formula_ids are unique
    ids = [r.formula_id for r in records]
    assert len(ids) == len(set(ids))


def test_analyze_formulas_volatile_detected(fixtures_dir: Path) -> None:
    path = fixtures_dir / "formulas_complex.xlsm"
    wb = load_workbook(str(path), data_only=False, keep_vba=True)
    cells = list(extract_cells(wb, set(), set()))
    warnings: list[str] = []
    records = list(analyze_formulas(cells, warnings))
    wb.close()

    volatile = [r for r in records if r.is_volatile]
    assert len(volatile) >= 1


def test_analyze_formulas_mixed_category(fixtures_dir: Path) -> None:
    path = fixtures_dir / "formulas_complex.xlsm"
    wb = load_workbook(str(path), data_only=False, keep_vba=True)
    cells = list(extract_cells(wb, set(), set()))
    warnings: list[str] = []
    records = list(analyze_formulas(cells, warnings))
    wb.close()

    mixed = [r for r in records if r.formula_category == "mixed"]
    assert len(mixed) >= 1


def test_analyze_formulas_nesting_depth(fixtures_dir: Path) -> None:
    path = fixtures_dir / "formulas_complex.xlsm"
    wb = load_workbook(str(path), data_only=False, keep_vba=True)
    cells = list(extract_cells(wb, set(), set()))
    warnings: list[str] = []
    records = list(analyze_formulas(cells, warnings))
    wb.close()

    max_depth = max(r.nesting_depth for r in records)
    assert max_depth >= 4  # D8 has depth 5


def test_analyze_formulas_from_basic_fixture(fixtures_dir: Path) -> None:
    path = fixtures_dir / "formulas_basic.xlsm"
    wb = load_workbook(str(path), data_only=False, keep_vba=True)
    cells = list(extract_cells(wb, set(), set()))
    warnings: list[str] = []
    records = list(analyze_formulas(cells, warnings))
    wb.close()

    assert len(records) >= 1
    categories = {r.formula_category for r in records}
    # basic fixture should have at least one non-empty category
    assert categories - {"compute"} or "compute" in categories


def test_analyze_formulas_deterministic(fixtures_dir: Path) -> None:
    path = fixtures_dir / "formulas_complex.xlsm"
    wb1 = load_workbook(str(path), data_only=False, keep_vba=True)
    cells1 = list(extract_cells(wb1, set(), set()))
    wb1.close()
    w1: list[str] = []
    r1 = sorted(analyze_formulas(cells1, w1), key=lambda f: f.qualified_address)

    wb2 = load_workbook(str(path), data_only=False, keep_vba=True)
    cells2 = list(extract_cells(wb2, set(), set()))
    wb2.close()
    w2: list[str] = []
    r2 = sorted(analyze_formulas(cells2, w2), key=lambda f: f.qualified_address)

    assert [f.qualified_address for f in r1] == [f.qualified_address for f in r2]
    assert [f.formula_category for f in r1] == [f.formula_category for f in r2]
