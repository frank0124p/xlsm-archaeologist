"""Phase 5 dependency graph tests."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from xlsm_archaeologist.analyzers.cycle_detector import detect_cycles
from xlsm_archaeologist.analyzers.dependency_analyzer import (
    graph_to_json,
    run_dependency_analysis,
)
from xlsm_archaeologist.analyzers.dependency_graph_builder import build_graph
from xlsm_archaeologist.analyzers.formula_analyzer import analyze_formulas
from xlsm_archaeologist.analyzers.orphan_detector import detect_orphans
from xlsm_archaeologist.extractors.cell_extractor import extract_cells
from xlsm_archaeologist.extractors.named_range_extractor import extract_named_ranges
from xlsm_archaeologist.extractors.validation_extractor import extract_validations
from xlsm_archaeologist.models.cell import CellRecord, ValueType
from xlsm_archaeologist.models.formula import CellRef, FormulaRecord


@pytest.fixture
def fixtures_dir() -> Path:
    return Path(__file__).parent / "fixtures"


def _make_formula(
    fid: int,
    qa: str,
    refs: list[str],
    named: list[str] | None = None,
) -> FormulaRecord:
    """Helper to build a minimal FormulaRecord."""
    cell_refs = []
    for r in refs:
        if "!" in r:
            sheet, addr = r.split("!", 1)
            cell_refs.append(CellRef(sheet=sheet, address=addr))
        else:
            cell_refs.append(CellRef(sheet=None, address=r))
    return FormulaRecord(
        formula_id=fid,
        qualified_address=qa,
        formula_text="=A1",
        formula_category="compute",
        function_list=[],
        referenced_cells=cell_refs,
        referenced_named_ranges=named or [],
        has_external_reference=False,
        is_volatile=False,
        is_array_formula=False,
        nesting_depth=0,
        function_count=0,
        complexity_score=0,
        ast=None,
        is_parsable=True,
        parse_error=None,
    )


def _make_cell(qa: str, has_formula: bool = False) -> CellRecord:
    vtype: ValueType = "number"
    return CellRecord(
        cell_id=1,
        sheet_name=qa.split("!")[0] if "!" in qa else "Sheet1",
        cell_address=qa.split("!")[1] if "!" in qa else qa,
        qualified_address=qa,
        cell_row=1,
        cell_col=1,
        has_formula=has_formula,
        has_validation=False,
        is_named=False,
        is_referenced=False,
        value_type=vtype,
        raw_value="",
    )


# ---------------------------------------------------------------------------
# Graph builder tests
# ---------------------------------------------------------------------------


def test_graph_basic_formula_dependency() -> None:
    formulas = [_make_formula(1, "Sheet1!B1", ["Sheet1!A1"])]
    cells = [_make_cell("Sheet1!A1"), _make_cell("Sheet1!B1", has_formula=True)]
    G = build_graph(formulas, [], [], cells, [])
    assert G.has_node("Sheet1!A1")
    assert G.has_node("Sheet1!B1")
    assert G.has_edge("Sheet1!A1", "Sheet1!B1")


def test_graph_named_range_node() -> None:
    formulas = [_make_formula(1, "Sheet1!B1", [], named=["TaxRate"])]
    cells = [_make_cell("Sheet1!B1", has_formula=True)]
    G = build_graph(formulas, [], [], cells, [])
    assert G.has_node("_named:TaxRate")
    assert G.has_edge("_named:TaxRate", "Sheet1!B1")


def test_graph_cross_sheet_marked() -> None:
    formulas = [_make_formula(1, "Sheet2!B1", ["Sheet1!A1"])]
    cells = [_make_cell("Sheet1!A1"), _make_cell("Sheet2!B1", has_formula=True)]
    G = build_graph(formulas, [], [], cells, [])
    edge_data = G.get_edge_data("Sheet1!A1", "Sheet2!B1")
    assert edge_data is not None
    assert edge_data["is_cross_sheet"] is True


# ---------------------------------------------------------------------------
# Cycle detection tests
# ---------------------------------------------------------------------------


def test_cycle_detection_simple() -> None:
    """A→B→A forms a 2-cycle."""
    import networkx as nx

    G: nx.DiGraph = nx.DiGraph()
    G.add_node("Sheet1!A1", node_type="formula_cell")
    G.add_node("Sheet1!B1", node_type="formula_cell")
    G.add_edge("Sheet1!A1", "Sheet1!B1", via="formula", via_detail="1")
    G.add_edge("Sheet1!B1", "Sheet1!A1", via="formula", via_detail="2")
    cycles = detect_cycles(G)
    assert len(cycles) >= 1
    assert cycles[0].length == 2


def test_cycle_detection_no_cycle() -> None:
    import networkx as nx

    G: nx.DiGraph = nx.DiGraph()
    G.add_edge("Sheet1!A1", "Sheet1!B1", via="formula", via_detail="1")
    G.add_edge("Sheet1!B1", "Sheet1!C1", via="formula", via_detail="2")
    cycles = detect_cycles(G)
    assert cycles == []


def test_cycle_from_circular_fixture(fixtures_dir: Path) -> None:
    path = fixtures_dir / "circular.xlsm"
    wb = load_workbook(str(path), data_only=False, keep_vba=True)
    cells = list(extract_cells(wb, set(), set()))
    wb.close()
    w: list[str] = []
    formulas = list(analyze_formulas(cells, w))
    G = build_graph(formulas, [], [], cells, [])
    cycles = detect_cycles(G)
    assert len(cycles) >= 1


# ---------------------------------------------------------------------------
# Orphan detection tests
# ---------------------------------------------------------------------------


def test_orphan_detection_finds_unreferenced_formula() -> None:
    import networkx as nx

    G: nx.DiGraph = nx.DiGraph()
    G.add_node("Sheet1!A1", node_type="input_cell")
    G.add_node("Sheet1!B1", node_type="formula_cell")  # no in-edges
    G.add_edge("Sheet1!B1", "Sheet1!A1", via="formula", via_detail="1")
    orphans = detect_orphans(G)
    assert "Sheet1!B1" in orphans


def test_orphan_detection_referenced_formula_not_orphan() -> None:
    import networkx as nx

    G: nx.DiGraph = nx.DiGraph()
    G.add_node("Sheet1!B1", node_type="formula_cell")
    G.add_node("Sheet1!C1", node_type="formula_cell")
    G.add_edge("Sheet1!A1", "Sheet1!B1", via="formula", via_detail="1")
    G.add_edge("Sheet1!B1", "Sheet1!C1", via="formula", via_detail="2")
    orphans = detect_orphans(G)
    assert "Sheet1!B1" not in orphans


# ---------------------------------------------------------------------------
# is_referenced backfill tests
# ---------------------------------------------------------------------------


def test_is_referenced_backfilled() -> None:
    formulas = [_make_formula(1, "Sheet1!B1", ["Sheet1!A1"])]
    input_cell = _make_cell("Sheet1!A1")
    formula_cell = _make_cell("Sheet1!B1", has_formula=True)
    cells = [input_cell, formula_cell]
    w: list[str] = []
    _, _, _, _, updated = run_dependency_analysis(formulas, [], [], cells, [], w)
    by_qa = {c.qualified_address: c for c in updated}
    # A1 is referenced by B1's formula
    assert by_qa["Sheet1!A1"].is_referenced is True
    # B1 is not referenced by anything
    assert by_qa["Sheet1!B1"].is_referenced is False


# ---------------------------------------------------------------------------
# Cross-sheet chain fixture
# ---------------------------------------------------------------------------


def test_cross_sheet_chain(fixtures_dir: Path) -> None:
    path = fixtures_dir / "cross_sheet_chain.xlsm"
    wb = load_workbook(str(path), data_only=False, keep_vba=True)
    cells = list(extract_cells(wb, set(), set()))
    wb.close()
    w: list[str] = []
    formulas = list(analyze_formulas(cells, w))
    G = build_graph(formulas, [], [], cells, [])
    cross_edges = [
        (s, t) for s, t, d in G.edges(data=True) if d.get("is_cross_sheet")
    ]
    assert len(cross_edges) >= 1


# ---------------------------------------------------------------------------
# Graph serialization
# ---------------------------------------------------------------------------


def test_graph_serialization_roundtrip() -> None:
    formulas = [_make_formula(1, "Sheet1!B1", ["Sheet1!A1"])]
    cells = [_make_cell("Sheet1!A1"), _make_cell("Sheet1!B1", has_formula=True)]
    G = build_graph(formulas, [], [], cells, [])
    data = graph_to_json(G, [])
    assert data["directed"] is True
    assert data["graph"]["node_count"] >= 2
    assert data["graph"]["edge_count"] >= 1
    node_ids = {n["id"] for n in data["nodes"]}  # type: ignore[index]
    assert "Sheet1!A1" in node_ids
    assert "Sheet1!B1" in node_ids
